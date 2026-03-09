from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.db.models.functions import TruncMonth
from django.utils import timezone
import json
import io

from activities.models import Activity, ActivityParticipation
from .models import Organization, OrganizationMember, Semester, AuditLog
from .decorators import admin_required, staff_required
from .permissions import can_manage_org_staff, can_create_org, get_manageable_orgs


@login_required
def dashboard_view(request):
    """Main dashboard with summary statistics. STUDENT gets redirected to their portal."""
    # Students should not see the admin/staff dashboard — send them to their own portal
    if request.user.role == 'STUDENT':
        return redirect('students:dashboard')

    from .permissions import get_point_category_orgs

    now = timezone.now()
    root_orgs = get_point_category_orgs(request.user)
    context = {
        'total_activities': Activity.objects.count(),
        'pending_count': Activity.objects.filter(status='PENDING').count(),
        'ongoing_count': Activity.objects.filter(status__in=['APPROVED', 'ONGOING']).count(),
        'total_organizations': Organization.objects.filter(status=True).count(),
        'current_semester': Semester.objects.filter(is_current=True, organization__in=root_orgs).first(),
        'recent_activities': Activity.objects.select_related('organization', 'created_by').order_by('-created_at')[:7],
    }
    return render(request, 'core/dashboard.html', context)


@staff_required
def unified_pending_view(request):
    """Unified center for all pending approvals across the system."""
    pending_activities_count = Activity.objects.filter(status='PENDING').count()
    pending_attendance_count = ActivityParticipation.objects.filter(status='ATTENDED').count()
    
    # Ready for future pending points or other modules.
    
    context = {
        'pending_activites_count': pending_activities_count,
        'pending_attendance_count': pending_attendance_count,
    }
    return render(request, 'core/unified_pending.html', context)

@login_required
def organizations_view(request):
    """
    List organizations.
    - ADMIN: sees all orgs in tree structure (parent -> children).
    - STAFF: sees only orgs they belong to (flat list).
    - Both: support search (q) and filter (org_type).
    """
    search_q = request.GET.get('q', '').strip()
    org_type_filter = request.GET.get('org_type', '').strip()

    # --- Base queryset ---
    base_qs = Organization.objects.filter(status=True).select_related('parent').annotate(
        member_count=Count('members')
    )

    # --- Role-based scoping ---
    is_admin = request.user.role == 'ADMIN'
    if not is_admin:
        # Staff: only orgs they are member of
        staff_org_ids = OrganizationMember.objects.filter(
            user=request.user
        ).values_list('organization_id', flat=True)
        base_qs = base_qs.filter(pk__in=staff_org_ids)

    # --- Search & Filter ---
    if search_q:
        base_qs = base_qs.filter(
            Q(name__icontains=search_q) | Q(code__icontains=search_q)
        )

    if org_type_filter:
        base_qs = base_qs.filter(type=org_type_filter)

    base_qs = base_qs.order_by('type', 'name')

    # --- Build tree for Admin (when not actively searching) ---
    org_tree = None
    if is_admin and not search_q and not org_type_filter:
        all_orgs = list(base_qs)
        # Group: root orgs (no parent or parent is inactive/missing)
        org_map = {o.pk: o for o in all_orgs}
        for o in all_orgs:
            o.tree_children = []
        roots = []
        for o in all_orgs:
            if o.parent_id and o.parent_id in org_map:
                org_map[o.parent_id].tree_children.append(o)
            else:
                roots.append(o)
        org_tree = roots  # nested via .tree_children

    context = {
        'organizations': base_qs,          # flat list (used by staff / when searching)
        'org_tree': org_tree,              # nested list (used by admin when not searching)
        'is_admin': is_admin,
        'is_searching': bool(search_q or org_type_filter),
        'search_q': search_q,
        'org_type_filter': org_type_filter,
        'org_type_choices': Organization.OrgType.choices,
        'can_create_org': can_create_org(request.user),
        'manageable_org_ids': set(
            get_manageable_orgs(request.user).values_list('id', flat=True)
        ),
    }
    return render(request, 'core/organizations.html', context)



@staff_required
def statistics_view(request):
    """
    Advanced statistics page with Chart.js data.
    Accessible by ADMIN and STAFF.
    """
    import datetime
    import io
    from django.http import HttpResponse

    # Dict maps for translations
    STATUS_MAP = dict(Activity.ActivityStatus.choices)
    TYPE_MAP = dict(Activity.ActivityType.choices)
    ORG_TYPE_MAP = dict(Organization.OrgType.choices)

    # --- Filtering ---
    semester_id = request.GET.get('semester')
    org_id = request.GET.get('organization')
    
    from .permissions import get_point_category_orgs, group_orgs_by_root
    root_orgs = get_point_category_orgs(request.user)
    
    if request.user.role != 'ADMIN':
        from .permissions import get_usable_point_category_orgs
        visible_orgs = get_usable_point_category_orgs(request.user)
        base_act_qs = Activity.objects.filter(organization__in=visible_orgs)
    else:
        base_act_qs = Activity.objects.all()

    if org_id:
        base_act_qs = base_act_qs.filter(organization_id=org_id)
        # Verify the org is within allowed root_orgs or admin
        if not (request.user.role == 'ADMIN' or int(org_id) in [o.id for o in root_orgs]):
            # Optional: handle unauthorized org access
            pass
            
    if semester_id:
        base_act_qs = base_act_qs.filter(semester_id=semester_id)
        
    org_groups = group_orgs_by_root(root_orgs)
    
    semesters = Semester.objects.filter(organization__in=root_orgs).order_by('-start_date')
    if org_id:
        # Provide only semesters belonging to the selected organization
        semesters = semesters.filter(organization_id=org_id)

    # ── Activity stats by status ────────────────────────────────────────────
    status_qs = (
        base_act_qs.values('status')
        .annotate(count=Count('id'))
        .order_by('status')
    )
    status_labels = [STATUS_MAP.get(s['status'], s['status']) for s in status_qs]
    status_data = [s['count'] for s in status_qs]

    # ── Activity stats by type ───────────────────────────────────────────────
    type_qs = (
        base_act_qs.values('activity_type')
        .annotate(count=Count('id'))
        .order_by('activity_type')
    )
    type_labels = [TYPE_MAP.get(t['activity_type'], t['activity_type']) for t in type_qs]
    type_data = [t['count'] for t in type_qs]

    # ── Monthly activity creation (last 6 months - padded) ───────────────────
    today = timezone.now()
    months_list = []
    current_month = today.month
    current_year = today.year
    for i in range(5, -1, -1):
        m = current_month - i
        y = current_year
        if m <= 0:
            m += 12
            y -= 1
        months_list.append(f"{m:02d}/{y}")

    six_months_ago = today - timezone.timedelta(days=180)
    monthly_qs = (
        base_act_qs
        .filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    monthly_counts = {m['month'].strftime('%m/%Y'): m['count'] for m in monthly_qs}
    monthly_labels = months_list
    monthly_data = [monthly_counts.get(m, 0) for m in months_list]

    # ── Budget stats ─────────────────────────────────────────────────────────
    budget_total = 0
    for act in base_act_qs:
        if act.budget_info and act.budget_info.get('status') == 'APPROVED':
            try:
                budget_total += float(act.budget_info.get('total_amount', 0))
            except (ValueError, TypeError):
                pass


    # ── Attendance stats ─────────────────────────────────────────────────────
    att_qs = ActivityParticipation.objects.filter(status='VERIFIED')
    if semester_id:
        att_qs = att_qs.filter(activity__semester_id=semester_id)
    total_checkins = att_qs.count()

    # ── Organization breakdown ───────────────────────────────────────────────
    org_qs = (
        Organization.objects
        .filter(status=True)
        .values('type')
        .annotate(count=Count('id'))
    )
    org_labels = [ORG_TYPE_MAP.get(o['type'], o['type']) for o in org_qs]
    org_data = [o['count'] for o in org_qs]
    
    # ── Export Logic ──
    if request.GET.get('export') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bao Cao Thong Ke'

        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = "BÁO CÁO THỐNG KÊ HOẠT ĐỘNG"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Summary
        headers = ['Tổng hoạt động', 'Chờ duyệt', 'Tổ chức', 'Điểm danh', 'Ngân sách (đ)']
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FF6A00', end_color='FF6A00', fill_type='solid')

        ws.append([
            base_act_qs.count(),
            base_act_qs.filter(status='PENDING').count(),
            Organization.objects.filter(status=True).count(),
            total_checkins,
            budget_total
        ])

        ws.append([]) # Empty row
        ws.append(['TRẠNG THÁI HOẠT ĐỘNG'])
        ws._current_row[-1][0].font = Font(bold=True)
        for i, lbl in enumerate(status_labels):
            ws.append([lbl, status_data[i]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="bao_cao_thong_ke.xlsx"'}
        )

    context = {
        # Chart.js data (JSON-safe)
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        'type_labels': json.dumps(type_labels),
        'type_data': json.dumps(type_data),
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'org_labels': json.dumps(org_labels),
        'org_data': json.dumps(org_data),
        # Summary figures
        'total_activities': base_act_qs.count(),
        'total_organizations': Organization.objects.filter(status=True).count(),
        'budget_total': budget_total,
        'total_checkins': total_checkins,
        'pending_count': base_act_qs.filter(status='PENDING').count(),
        # Filter options
        'semesters': semesters,
        'current_semester_id': semester_id,
        'org_groups': org_groups,
        'current_org': org_id,
    }
    return render(request, 'core/statistics.html', context)


@admin_required
def audit_log_view(request):
    """View audit logs. ADMIN only."""
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')

    # Filter by action
    action_filter = request.GET.get('action')
    if action_filter:
        logs = logs.filter(action=action_filter)

    # Filter by object type
    obj_type = request.GET.get('type')
    if obj_type:
        logs = logs.filter(object_type__icontains=obj_type)

    # Search by user
    user_q = request.GET.get('user')
    if user_q:
        logs = logs.filter(user__full_name__icontains=user_q)

    type_choices = AuditLog.objects.exclude(object_type__isnull=True).exclude(object_type='').values_list('object_type', flat=True).distinct().order_by('object_type')

    context = {
        'logs': logs[:200],  # Limit to last 200
        'action_choices': AuditLog.Action.choices,
        'type_choices': type_choices,
        'current_action': action_filter or '',
        'current_type': obj_type or '',
        'current_user': user_q or '',
    }
    return render(request, 'core/audit_log.html', context)


# ─── B4: Create Organization (Admin Only) ─────────────────────────────────────

@admin_required
def organization_create(request):
    """B4: Create a new organization. ADMIN only."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        org_type = request.POST.get('type', '')
        parent_id = request.POST.get('parent') or None
        description = request.POST.get('description', '').strip()

        if not name or not code or not org_type:
            messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc.')
        elif Organization.objects.filter(code__iexact=code).exists():
            messages.error(request, 'Mã tổ chức đã tồn tại.')
        else:
            org = Organization.objects.create(
                name=name,
                code=code.upper(),
                type=org_type,
                parent_id=parent_id,
                description=description,
                status=True,
            )
            messages.success(request, f'Đã tạo tổ chức "{org.name}" thành công!')
            return redirect('core:organizations')

    context = {
        'parent_orgs': Organization.objects.filter(status=True).order_by('type', 'name'),
        'type_choices': Organization.OrgType.choices,
    }
    return render(request, 'core/organization_form.html', context)


# ─── B4b: Edit Organization (Admin Only) ──────────────────────────────────────

@admin_required
def organization_edit(request, org_pk):
    """Edit an existing organization. ADMIN only."""
    org = get_object_or_404(Organization, pk=org_pk)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        org_type = request.POST.get('type', '')
        parent_id = request.POST.get('parent') or None
        description = request.POST.get('description', '').strip()

        if not name or not code or not org_type:
            messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc.')
        elif Organization.objects.filter(code__iexact=code).exclude(pk=org.pk).exists():
            messages.error(request, 'Mã tổ chức đã tồn tại.')
        else:
            org.name = name
            org.code = code.upper()
            org.type = org_type
            org.parent_id = parent_id
            org.description = description
            org.save()
            messages.success(request, f'Đã cập nhật tổ chức "{org.name}" thành công!')
            return redirect('core:organizations')

    context = {
        'org': org,
        'parent_orgs': Organization.objects.filter(status=True).exclude(pk=org.pk).order_by('type', 'name'),
        'type_choices': Organization.OrgType.choices,
        'is_edit': True,
    }
    return render(request, 'core/organization_form.html', context)


# ─── B4c: Delete Organization (Admin Only) ─────────────────────────────────────

def _delete_org_tree(org):
    """Recursively delete an org and all its descendants."""
    for child in org.children.all():
        _delete_org_tree(child)
    org.delete()


@admin_required
def organization_delete(request, org_pk):
    """Delete an organization and all its children. ADMIN only."""
    org = get_object_or_404(Organization, pk=org_pk)

    if request.method == 'POST':
        org_name = org.name
        _delete_org_tree(org)
        messages.success(request, f'Đã xóa tổ chức "{org_name}" và tất cả tổ chức con.')
        return redirect('core:organizations')

    # Count children recursively for confirmation display
    from core.permissions import get_all_child_orgs
    children = get_all_child_orgs(org)
    member_count = OrganizationMember.objects.filter(organization=org).count()

    context = {
        'org': org,
        'children_count': len(children),
        'member_count': member_count,
    }
    return render(request, 'core/org_confirm_delete.html', context)


# ─── B3: Manage Members of an Org ────────────────────────────────────────────

@login_required
def org_staff_view(request, org_pk):
    """
    B3: View and manage ALL members of an org (Officers + Students).
    Now shows two split tables: Cán bộ and Đoàn viên/Hội viên.
    """
    org = get_object_or_404(Organization, pk=org_pk, status=True)

    if not can_manage_org_staff(request.user, org):
        messages.error(request, 'Bạn không có quyền quản lý thành viên của tổ chức này.')
        return redirect('core:organizations')

    from django.contrib.auth import get_user_model
    from users.models import StudentProfile
    User = get_user_model()

    all_members = OrganizationMember.objects.filter(
        organization=org
    ).select_related('user').order_by('user__full_name')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_officer':
            user_id = request.POST.get('user_id')
            position = request.POST.get('position', 'Can bo').strip()
            try:
                target_user = User.objects.get(pk=user_id, role='STAFF')
                member, created = OrganizationMember.objects.get_or_create(
                    organization=org,
                    user=target_user,
                    defaults={
                        'position': position,
                        'is_officer': True,
                        'joined_at': timezone.now().date(),
                    }
                )
                if not created:
                    member.is_officer = True
                    member.position = position
                    member.save(update_fields=['is_officer', 'position'])
                messages.success(request, f'Đã thêm "{target_user.full_name}" làm cán bộ.')
            except User.DoesNotExist:
                messages.error(request, 'Không tìm thấy user hoặc user không phải Staff.')

        elif action == 'add_student':
            student_email = request.POST.get('student_email', '').strip()
            if not student_email:
                messages.error(request, 'Vui lòng nhập Email.')
            else:
                target_user = User.objects.filter(email__iexact=student_email).first()
                if not target_user:
                    messages.error(request, f'Không tìm thấy thành viên có Email "{student_email}".')
                else:
                    member, created = OrganizationMember.objects.get_or_create(
                        organization=org,
                        user=target_user,
                        defaults={
                            'position': 'Thanh vien',
                            'is_officer': False,
                            'joined_at': timezone.now().date(),
                        }
                    )
                    if created:
                        messages.success(request, f'Đã thêm "{target_user.full_name}" vào tổ chức.')
                    else:
                        messages.info(request, f'"{target_user.full_name}" đã là thành viên của tổ chức rồi.')

        elif action == 'remove_officer':
            member_id = request.POST.get('member_id')
            member = OrganizationMember.objects.filter(pk=member_id, organization=org).first()
            if member:
                member.is_officer = False
                member.save(update_fields=['is_officer'])
                messages.success(request, f'Đã xoá quyền cán bộ khỏi "{member.user.full_name}".')

        elif action == 'remove_member':
            member_id = request.POST.get('member_id')
            member = OrganizationMember.objects.filter(pk=member_id, organization=org).first()
            if member:
                member.delete()
                messages.success(request, 'Đã xoá thành viên khỏi tổ chức.')

        return redirect('core:org_staff', org_pk=org_pk)

    # Split into two groups
    officers = all_members.filter(is_officer=True)
    regular_members = all_members.filter(is_officer=False)

    # Available staff users not yet in this org
    existing_user_ids = all_members.values_list('user_id', flat=True)
    available_staff = User.objects.filter(role='STAFF', status='ACTIVE').exclude(
        pk__in=existing_user_ids
    ).order_by('full_name')

    context = {
        'org': org,
        'officers': officers,
        'regular_members': regular_members,
        'all_members': all_members,
        'available_staff': available_staff,
    }
    return render(request, 'core/org_staff.html', context)


# ─── B5b: Import Members into a specific Org ─────────────────────────────────

@staff_required
def import_members_to_org(request, org_pk):
    """
    Import students from Excel directly into a specific organization.
    All imported students become members of THIS org, regardless of
    what faculty/class columns say in the Excel file.
    """
    from django.contrib.auth import get_user_model
    from users.models import StudentProfile

    User = get_user_model()

    org = get_object_or_404(Organization, pk=org_pk, status=True)

    # Permission check
    if not can_manage_org_staff(request.user, org):
        messages.error(request, 'Bạn không có quyền import thành viên vào tổ chức này.')
        return redirect('core:org_staff', org_pk=org_pk)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl

        excel_file = request.FILES['excel_file']
        default_password = request.POST.get('default_password', '').strip()

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'File không hợp lệ. Vui lòng tải lên file .xlsx đúng định dạng.')
            return redirect('core:org_staff', org_pk=org_pk)

        rows_ok, rows_skip, rows_error = [], [], []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            full_name = str(row[0]).strip() if row[0] else ''
            student_code = str(row[1]).strip() if row[1] else ''
            email = str(row[2]).strip() if row[2] else ''
            # Columns D,E,F used for profile info only, not org creation
            faculty_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            class_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            course_year = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            password_col = str(row[6]).strip() if len(row) > 6 and row[6] else ''

            if not full_name and not student_code and not email:
                continue

            if not full_name or not student_code or not email:
                rows_error.append(f'Dong {row_idx}: Thieu thong tin bat buoc (Ten/MSSV/Email).')
                continue

            # Check if user already exists
            existing_user = None
            profile = StudentProfile.objects.filter(student_code__iexact=student_code).first()
            if profile:
                existing_user = profile.user
            elif User.objects.filter(email__iexact=email).exists():
                rows_skip.append(f'Dong {row_idx}: Email "{email}" da ton tai nhung MSSV khong khop.')
                continue

            if existing_user:
                # User exists: just add to org
                member, created = OrganizationMember.objects.get_or_create(
                    organization=org,
                    user=existing_user,
                    defaults={
                        'position': 'Thanh vien',
                        'is_officer': False,
                        'joined_at': timezone.now().date(),
                    }
                )
                if created:
                    rows_ok.append(f'{student_code} — {full_name} (da co tai khoan, them vao to chuc)')
                else:
                    rows_skip.append(f'Dong {row_idx}: {student_code} da la thanh vien.')
            else:
                # Create new user + profile + member
                username = student_code.lower().replace(' ', '')
                password = password_col or default_password or student_code

                try:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        full_name=full_name,
                        role='STUDENT',
                        status='ACTIVE',
                    )
                    StudentProfile.objects.create(
                        user=user,
                        student_code=student_code,
                        course_year=course_year,
                    )
                    OrganizationMember.objects.create(
                        organization=org,
                        user=user,
                        position='Thanh vien',
                        is_officer=False,
                        joined_at=timezone.now().date(),
                    )
                    rows_ok.append(f'{student_code} — {full_name}')
                except Exception as e:
                    rows_error.append(f'Dong {row_idx}: Loi tao tai khoan — {e}')
                    continue

        if rows_ok:
            messages.success(request, f'Import thành công {len(rows_ok)} sinh viên vào "{org.name}".')
        if rows_skip:
            messages.warning(request, f'Bỏ qua {len(rows_skip)} dòng trùng lặp.')
        if rows_error:
            messages.error(request, f'Có {len(rows_error)} dòng lỗi.')

    return redirect('core:org_staff', org_pk=org_pk)


# ─── B5: Import Students from Excel ──────────────────────────────────────────

@staff_required
def import_students_view(request):
    """
    B5: Bulk-import student accounts from an Excel file.

    Excel columns (row 1 = header, row 2+ = data):
        A: full_name      (required)
        B: student_code   (required, unique)
        C: email          (required, unique)
        D: faculty_name   (required — matched/created as UNION_FACULTY org)
        E: class_name     (optional — matched/created as CLASS org under faculty)
        F: course_year    (optional)
        G: password       (optional — default: student_code)

    For each row the view will:
        1. get_or_create the UNION_SCHOOL org (first one found)
        2. get_or_create UNION_FACULTY org with parent = school
        3. get_or_create CLASS org with parent = faculty (if class_name provided)
        4. create User (role=STUDENT)
        5. create StudentProfile
        6. add OrganizationMember for school, faculty, and class (if any)
    """
    from django.contrib.auth import get_user_model
    from users.models import StudentProfile

    User = get_user_model()

    is_admin = request.user.role == 'ADMIN'
    manageable_orgs = get_manageable_orgs(request.user)

    results = None  # will be set after processing

    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl

        excel_file = request.FILES['excel_file']
        default_password = request.POST.get('default_password', '').strip()

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'File không hợp lệ. Vui lòng tải lên file .xlsx đúng định dạng.')
            return redirect('core:import_students')

        # Find the school org once
        school_org = Organization.objects.filter(
            type=Organization.OrgType.UNION_SCHOOL, status=True
        ).first()

        rows_ok, rows_skip, rows_error = [], [], []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Unpack columns
            full_name   = str(row[0]).strip() if row[0] else ''
            student_code = str(row[1]).strip() if row[1] else ''
            email       = str(row[2]).strip() if row[2] else ''
            faculty_name = str(row[3]).strip() if row[3] else ''
            class_name  = str(row[4]).strip() if row[4] else ''
            course_year = str(row[5]).strip() if row[5] else ''
            password_col = str(row[6]).strip() if len(row) > 6 and row[6] else ''

            # Skip blank rows
            if not full_name and not student_code and not email:
                continue

            # Validate required fields
            if not full_name or not student_code or not email or not faculty_name:
                rows_error.append(f'Dòng {row_idx}: Thiếu thông tin bắt buộc (Tên/MSSV/Email/Khoa).')
                continue

            # Skip duplicates
            if User.objects.filter(email__iexact=email).exists():
                rows_skip.append(f'Dòng {row_idx}: Email "{email}" đã tồn tại.')
                continue
            if StudentProfile.objects.filter(student_code__iexact=student_code).exists():
                rows_skip.append(f'Dòng {row_idx}: MSSV "{student_code}" đã tồn tại.')
                continue

            # ── Resolve / create Organizations ────────────────────────────────

            # Faculty org: UNION_FACULTY, parent = school
            faculty_code = 'FAC-' + faculty_name.upper().replace(' ', '-')[:20]
            faculty_org, _ = Organization.objects.get_or_create(
                type=Organization.OrgType.UNION_FACULTY,
                name=faculty_name,
                defaults={
                    'code': faculty_code,
                    'parent': school_org,
                    'status': True,
                }
            )

            # --- AuthZ Check for Staff ---
            if not is_admin and faculty_org not in manageable_orgs:
                rows_error.append(f"Dòng {row_idx}: Bạn không có quyền quản lý Khoa '{faculty_name}'. Import bị từ chối.")
                continue
            # -----------------------------

            # Class org (Chi đoàn): CLASS, parent = faculty
            class_org = None
            if class_name:
                class_code = 'CLS-' + class_name.upper().replace(' ', '-')[:20]
                class_org, _ = Organization.objects.get_or_create(
                    type=Organization.OrgType.CLASS,
                    name=class_name,
                    defaults={
                        'code': class_code,
                        'parent': faculty_org,
                        'status': True,
                    }
                )

            # ── Create User & StudentProfile ───────────────────────────────────
            # Auto-generate username from student_code (lowercase, no spaces)
            username = student_code.lower().replace(' ', '')
            # Auto-generate password
            password = password_col or default_password or student_code

            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    full_name=full_name,
                    role='STUDENT',
                    status='ACTIVE',
                )
                StudentProfile.objects.create(
                    user=user,
                    student_code=student_code,
                    course_year=course_year,
                )
            except Exception as e:
                rows_error.append(f'Dòng {row_idx}: Lỗi tạo tài khoản — {e}')
                continue

            # ── Add OrganizationMember ─────────────────────────────────────────
            joined_today = timezone.now().date()
            if school_org:
                OrganizationMember.objects.get_or_create(
                    organization=school_org, user=user,
                    defaults={'position': 'Đoàn viên', 'is_officer': False, 'joined_at': joined_today}
                )
            OrganizationMember.objects.get_or_create(
                organization=faculty_org, user=user,
                defaults={'position': 'Đoàn viên', 'is_officer': False, 'joined_at': joined_today}
            )
            if class_org:
                OrganizationMember.objects.get_or_create(
                    organization=class_org, user=user,
                    defaults={'position': 'Đoàn viên', 'is_officer': False, 'joined_at': joined_today}
                )

            rows_ok.append(f'{student_code} — {full_name} ({faculty_name})')

        results = {
            'ok': rows_ok,
            'skip': rows_skip,
            'error': rows_error,
        }
        if rows_ok:
            messages.success(request, f'Import thành công {len(rows_ok)} sinh viên.')
        if rows_skip:
            messages.warning(request, f'Bỏ qua {len(rows_skip)} dòng trùng lặp.')
        if rows_error:
            messages.error(request, f'Có {len(rows_error)} dòng lỗi.')

    context = {
        'results': results,
        'school_org': Organization.objects.filter(
            type=Organization.OrgType.UNION_SCHOOL, status=True
        ).first(),
    }
    return render(request, 'core/import_students.html', context)


@staff_required
def download_import_template(request):
    """Return an Excel template for student import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Sinh Vien'

    headers = [
        'Họ và Tên (*)',
        'Mã sinh viên (*)',
        'Email (*)',
        'Tên Khoa (*)',
        'Tên Lớp',
        'Khóa học (VD: 2023)',
        'Mật khẩu (để trống = dùng MSSV)',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6A00', end_color='FF6A00', fill_type='solid')

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data rows
    samples = [
        ('Nguyễn Văn An', 'DNC2023001', 'an.nguyen@dnc.edu.vn', 'Khoa CNTT', 'CNTT-2023A', '2023', ''),
        ('Trần Thị Bình', 'DNC2023002', 'binh.tran@dnc.edu.vn', 'Khoa Kinh Tế', 'KT-2023B', '2023', ''),
        ('Lê Minh Cường', 'DNC2023003', 'cuong.le@dnc.edu.vn', 'Khoa CNTT', 'CNTT-2023A', '2023', 'mypassword'),
    ]
    for row_data in samples:
        ws.append(row_data)

    # Column widths
    col_widths = [25, 18, 30, 20, 18, 18, 25]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="template_import_sinhvien.xlsx"'
    return response

